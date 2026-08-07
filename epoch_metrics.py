"""逐-epoch 生理指標：input_folder 的整理版 xlsx。

來源：
    …/我的筆記型電腦/input_folder/{normalweight|overweight}/{溫度}/{場次}.xlsx

每個檔一張工作表，左半是 RemLogic 判讀（期別、時鐘），右半是廠商軟體算好的
每 30 秒指標，兩邊已經逐列對齊：

    Scoring Time: │ 2022/10/16 … │ Event │ │ .\\file.txt │ No. │ Time │ Stage │ MPF │ RR │ n │ SD │ RMSSD │ TP │ …
    SLEEP-S0      │ PM 11:11:14  │ W     │ │             │ 1   │23:11:14│  1   │4.275│1178│19 │104.3│ 125.3 │8.86│ …

用廠商算好的數值而非自行重算，是為了與既有分析維持同一把尺
（廠商與自算的 LF/HF 實測差約 4.9 倍，兩者不可混用）。

檔案裡含有受測者姓名，本模組不讀取、不顯示那一格。
"""
from __future__ import annotations

import datetime as dt
import glob
import json
import os
import re

import numpy as np

EPOCH = 30.0

INPUT_ROOT = ("/Users/appletina/Library/CloudStorage/GoogleDrive-mascottofu@gmail.com/"
              "其他電腦/我的筆記型電腦/input_folder")
CACHE = os.path.expanduser("~/.edf_viewer_cache/input_index.json")

HDR_ROW = 1          # 第 1 列是表頭
COL_NO = 6           # 'No.' 在第 6 欄（1-based）

# 這些欄位是 ln 值（表頭沒有單位列，依廠商 use.bpp 的定義）
LN_MS2 = {"TP", "HF", "LF", "VLF"}
LN_UV2 = {"delta", "theta", "alpha", "beta", "alleeg", "Sigma", "EOG", "EMG"}
UNITS = {"MPF": "Hz", "RR": "ms", "n": "beats", "SD": "ms", "RMSSD": "ms",
         "LF/HF": "ratio", "LF%": "%", "d%": "%", "t%": "%", "a%": "%", "b%": "%",
         "Sigma%": "%", "PA": "", "PHI": "r", "THETA": "r", "Stage": "code"}

NOTES = {
    "MPF": "平均頻率（mean power frequency）",
    "RR": "心搏間期",
    "n": "該格偵測到的心搏數",
    "SD": "RR 間期標準差（SDNN）",
    "RMSSD": "相鄰 RR 差值的均方根",
    "TP": "總功率 0.003–0.4 Hz",
    "HF": "高頻 0.15–0.4 Hz",
    "LF": "低頻 0.04–0.15 Hz",
    "VLF": "極低頻 0.003–0.04 Hz",
    "LF/HF": "LF 與 HF 的比值（不等於交感／副交感平衡）",
    "LF%": "LF / (LF+HF) × 100",
    "delta": "0.5–4 Hz", "theta": "4–8 Hz", "alpha": "8–13 Hz", "beta": "13–32 Hz",
    "Sigma": "12–14 Hz（與 alpha/beta 重疊）",
    "alleeg": "全頻段總功率",
    "d%": "delta 佔比", "t%": "theta 佔比", "a%": "alpha 佔比", "b%": "beta 佔比",
    "Sigma%": "sigma 佔比",
    "EMG": "肌電 RMS", "EOG": "眼動",
    "Stage": "廠商軟體的期別代碼",
}

_DIGITS = re.compile(r"\d{6,}")


def _tokens(name):
    """檔名裡長度 ≥6 的數字串（日期碼）。"""
    return {m.group(0) for m in _DIGITS.finditer(name)}


def _token_match(edf_stem, xlsx_stem):
    """xlsx 檔名常帶受測者前綴（27170406h16 = GT027 + 170406h + 16°C），
    所以比對用「包含」而非相等：EDF 的日期碼要出現在 xlsx 的數字串裡。"""
    a, b = _tokens(edf_stem), _tokens(xlsx_stem)
    return any(x in y or y in x for x in a for y in b)


class EpochMetrics:
    def __init__(self, path, rows, cols, first_clock, offset_epochs):
        self.path = path
        self.source = os.path.relpath(path, INPUT_ROOT)
        self.columns = cols
        self.first_clock = first_clock            # 第 1 格的時鐘時間
        self.offset = offset_epochs               # 指標第 1 格相對記錄起點的格數
        self._rows = rows                         # {檔內格號: {欄位: float|None}}
        self.n = (max(rows) + 1) if rows else 0
        self.match_reason = ""

    def at(self, t):
        """記錄開始後 t 秒落在哪一格。"""
        return self._rows.get(int(t // EPOCH) - self.offset)

    def series(self, col):
        """對齊到記錄起點的整段序列，缺值為 nan。"""
        out = np.full(self.n + self.offset, np.nan)
        for i, r in self._rows.items():
            v = r.get(col)
            if v is not None and 0 <= i + self.offset < len(out):
                out[i + self.offset] = v
        return out

    def is_ln(self, col):
        return col in LN_MS2 or col in LN_UV2

    def unit(self, col):
        if col in LN_MS2:
            return "ln(ms²)"
        if col in LN_UV2:
            return "ln(µV²)"
        return UNITS.get(col, "")

    def note(self, col):
        return NOTES.get(col, "")


# ---------------------------------------------------------------- 索引


def _layout(hdr, first):
    """找出「格號欄」與指標欄名。

    表頭不統一 —— 有些檔的 'No.' 那格被檔名佔掉了。所以改用結構定位：
    從 'MPF' 往左跳過 Stage / Time 這兩個固定欄，再左邊那格就是格號欄。
    （不能用「第一列的值等於 1」來找 —— Stage 欄的值也是 1，會選錯。）
    """
    hs = [str(h).strip() if h is not None else "" for h in hdr]
    try:
        mpf = hs.index("MPF")
    except ValueError:
        raise ValueError("表頭裡找不到 'MPF' 欄")
    i = mpf - 1
    while i >= 0 and hs[i] in ("Stage", "Time"):
        i -= 1
    if i < 0:
        raise ValueError("找不到 epoch 格號欄")
    if not isinstance(first[i], (int, float)):
        raise ValueError("格號欄的第一列不是數字")
    return i, hs[i + 1:]


def _scan_one(path):
    """只讀表頭與第一列資料，取出配對用的資訊（不碰含姓名的那一格）。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(min_row=1, max_row=2, values_only=True)
    hdr = next(it, None)
    first = next(it, None)
    title = ws.title
    wb.close()
    if not hdr or not first:
        return None
    try:
        ecol, names = _layout(hdr, first)
        clock = first[ecol + 1] if names and names[0] == "Time" else None
    except Exception:
        return None
    if isinstance(clock, dt.time):
        clock = clock.strftime("%H:%M:%S")
    elif clock is not None:
        clock = str(clock)
    rel = os.path.relpath(path, INPUT_ROOT)
    parts = rel.split(os.sep)
    return {
        "path": path,
        "group": parts[0] if parts else "",
        "temp": parts[1] if len(parts) > 2 else "",
        "stem": os.path.splitext(os.path.basename(path))[0],
        "sheet": title,
        "clock": clock,
    }


def build_index(progress=None):
    files = sorted(glob.glob(os.path.join(INPUT_ROOT, "**", "*.xlsx"), recursive=True))
    idx = []
    for k, p in enumerate(files):
        if os.path.basename(p).startswith("~$"):
            continue
        try:
            e = _scan_one(p)
        except Exception:
            e = None
        if e:
            idx.append(e)
        if progress:
            progress((k + 1) / len(files))
    os.makedirs(os.path.dirname(CACHE), exist_ok=True)
    json.dump(idx, open(CACHE, "w"), ensure_ascii=False)
    return idx


def load_index():
    if os.path.exists(CACHE):
        try:
            return json.load(open(CACHE))
        except Exception:
            pass
    return None


# ---------------------------------------------------------------- 配對


def _clock_secs(s):
    try:
        h, m, sec = (int(x) for x in str(s).split(":")[:3])
        return h * 3600 + m * 60 + sec
    except Exception:
        return None


def match(edf_name, start_datetime, index):
    """把一個 EDF 對到 input_folder 裡的 xlsx。回傳 (entry|None, 理由)。

    `data/` 的檔名已經帶了組別與溫度（NW_GT028_P2_28C_…），加上場次代號與
    起始時刻，四個條件互相佐證。沒有正面證據就回報配不到，不猜 —— 一個溫度
    資料夾裡放了多位受測者，猜錯會把別人的生理數據貼到這場記錄上。
    """
    if not index:
        return None, "尚未建立索引"

    stem = os.path.splitext(os.path.basename(edf_name))[0]
    grp = {"NW": "normalweight", "OW": "overweight"}.get(stem.split("_")[0])
    m = re.search(r"_(\d{2})C_", stem)
    temp = m.group(1) if m else None
    edf_secs = (start_datetime.hour * 3600 + start_datetime.minute * 60
                + start_datetime.second) if start_datetime else None

    best = None
    for e in index:
        tok = _token_match(stem, e["stem"])
        grp_ok = (grp is None) or (e["group"] == grp)
        temp_ok = (temp is None) or (e["temp"] == temp) or (e["temp"].lstrip("0") == temp.lstrip("0"))
        clk = _clock_secs(e["clock"])
        # 指標第 1 格通常晚於記錄起點（熄燈前那段沒判讀），允許 0–2 小時的正向差
        delta = None
        if clk is not None and edf_secs is not None:
            delta = (clk - edf_secs) % 86400
            clock_ok = delta <= 7200
        else:
            clock_ok = False
        if not grp_ok or not temp_ok:
            continue
        # 場次代號是必要條件。時鐘只能當佐證 —— 同一個溫度資料夾裡有多位
        # 受測者，熄燈時間本來就接近，光靠時刻會把別人的生理數據配上來
        # （實測 GT031 的 8 晚就會全部配到 GT028 身上）。
        if not tok:
            continue
        score = 4 + 2 * clock_ok
        key = (score, -(delta if delta is not None else 9e9))
        if best is None or key > best[0]:
            best = (key, tok, clock_ok, delta, e)
    if best is None:
        return None, (f"input_folder 裡找不到場次代號相符的檔"
                      f"（組別 {grp}／溫度 {temp}）")

    _, tok, clock_ok, delta, e = best
    why = []
    if tok:
        why.append("場次代號相符")
    if clock_ok:
        why.append(f"起始時刻相差 {int(delta // 60)} 分")
    return e, "、".join(why)


# ---------------------------------------------------------------- 讀檔


def parse(path, offset_epochs=0):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    first = next(it)
    ecol, names = _layout(hdr, first)
    cols = [n for n in names if n and n != "Time"]

    rows, first_clock = {}, None
    for r in (first, *it):
        if len(r) <= ecol or not isinstance(r[ecol], (int, float)):
            continue
        i = int(r[ecol]) - 1
        vals = {}
        for k, name in enumerate(names):
            v = r[ecol + 1 + k] if ecol + 1 + k < len(r) else None
            if name == "Time":
                if first_clock is None and isinstance(v, dt.time):
                    first_clock = v.strftime("%H:%M:%S")
                continue
            if name:
                vals[name] = float(v) if isinstance(v, (int, float)) else None
        rows[i] = vals
    wb.close()
    return EpochMetrics(path, rows, cols, first_clock, offset_epochs)


def load(edf_path, start_datetime=None, index=None):
    """回傳 (EpochMetrics|None, 說明)。"""
    index = index if index is not None else load_index()
    if index is None:
        return None, "尚未建立 input_folder 索引（側欄按「重建索引」）"
    e, why = match(os.path.basename(edf_path), start_datetime, index)
    if e is None:
        return None, why
    off = 0
    clk = _clock_secs(e["clock"])
    if clk is not None and start_datetime is not None:
        edf_secs = (start_datetime.hour * 3600 + start_datetime.minute * 60
                    + start_datetime.second)
        off = int(round(((clk - edf_secs) % 86400) / EPOCH))
    m = parse(e["path"], off)
    m.match_reason = why
    return m, why
